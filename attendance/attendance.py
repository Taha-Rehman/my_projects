from __future__ import print_function, unicode_literals
from loadstudents import *
from datetime import date
import datetime
from datetime import date, timedelta
import os
from PyInquirer import style_from_dict, Token, prompt, Separator
from pprint import pprint
from optparse import OptionParser
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import sys

parser = OptionParser()
parser.add_option("-a", "--attendance", dest="studentattendance", help="it takes attendance")
parser.add_option("-s", "--student", dest="student", help="Enter student")
parser.add_option("-r","--removestudent", dest="removestudent", help="Enter rollnumber to delete")
parser.add_option("-k", "--knowone", dest="knowone", help="Enter enter student rollnumber to get related data")

(options, arguments) = parser.parse_args()


def attendance():
	path = "students.xlsx"
	# loading workbook
	wb = openpyxl.load_workbook(path) 
	# activate sheet
	sheet = wb.active 
	# getting the firsy column
	row = sheet.max_row
	row_a = []
	for i in range(1, row + 1): 
		cell_obj = sheet.cell(row = i, column = 1) 
		row_a.append(cell_obj.value)
		
	# getting the last element
	last_rn = row_a[-1]
	# make sure it is an integer
	if isinstance(last_rn, int):
		pass
	else:
		for i in row_a[::-1]:
			if isinstance(i, int) == False:
				row -= 1

	cell_obj = sheet['A2': 'B'+str(row+1)]

	students_lst = []
	for cell1, cell2 in cell_obj:

		students_lst.append(str(cell1.value) + " " + cell2.value)
	wb.close()
	# making interface for attendance
	style = style_from_dict({
		Token.QuestionMark: '#E91E63 bold',
		Token.Selected: '#673AB7 bold',
		Token.Instruction: '',  # default
		Token.Answer: '#2196f3 bold',
		Token.Question: '',
	})
	# preparing students for attendance
	attendance = []
	for i in students_lst:
		attendance.append({
			'type': 'confirm',
			'name': i,
			'message': i,
			'default': True
		})

	# taking attendance
	global answers
	answers = prompt(attendance, style=style)
	return answers


def get_weekdays():
	# -------------
	# from here to...
	d = str(date.today()).split("-")
	start_date = "/".join([d[1], d[2], d[0][-2:]])
	date_1 = datetime.datetime.strptime(start_date, "%m/%d/%y")
	end_date = date_1 + datetime.timedelta(days=30)
	startdate = [int(i) for i in (str(date_1)[:-9]).split("-")]
	end_date = [int(i) for i in (str(end_date)[:-9]).split("-")]
	# here, its a bad code :)...just getting the start and end date of the month

	test_date1, test_date2 = datetime.datetime(startdate[0], startdate[1], startdate[2]), datetime.datetime(end_date[0], end_date[1], end_date[2])
	
	# generating dates
	dates = (test_date1 + timedelta(idx + 1)
			for idx in range((test_date2 - test_date1).days))
	
	# summing all weekdays
	res = sum(1 for day in dates if day.weekday() < 5)
	return int(res)

def load_attendance_into_excel():
	weekdays = get_weekdays()
	if len(os.listdir("attendance")) < 1:
		print("[-] No file found to analyze")
		print("[+] Creating \"1.xlsx\"")
		newbook = "1.xlsx"
		newbook = openpyxl.Workbook(newbookname)
		newbook.save("attendance\\"+newbookname)

	file = "attendance\\"+os.listdir("attendance")[-1]
	wb = openpyxl.load_workbook(file)
	# sheet names
	sheet_names = wb.sheetnames
	# global info
	
	if len(sheet_names) == 1:
		budget = int(input("Enter Budget: "))
		ss_sheet = wb[sheet_names[-1]]
		ss_sheet.title = "1"
		wb.create_sheet('info')
		info = wb["info"]
		info["A1"].value = "Budget"
		info["B1"].value = "Rs."+str(budget)
		info["A2"].value = "Working Days"
		info.column_dimensions['A'].width = 30
		info["B2"].value = weekdays
		info["A3"].value = "Roll no."
		info["B3"].value = "Name"
		info.column_dimensions['B'].width = 70
		info["C3"].value = "Salary"

	# loading attendence data into

	info = wb["info"]
	weekdays = info["B2"].value
	sheet_names = wb.sheetnames
	del sheet_names[sheet_names.index("info")]
	sheet = wb[sheet_names[-1]]
	sheet["A1"].value = "-".join((str(date.today())).split("-")[::-1])
	sheet["A2"].value = "Roll No."
	sheet["B2"].value = "Name"
	sheet["C2"].value = "Attendance"
	sheet.column_dimensions['B'].width = 70
	sheet.column_dimensions['C'].width = 30
	sheet.column_dimensions['A'].width = 30
	student_money = {}
	for i, answer in enumerate(answers.keys()):
		index = str(i + 4)
		sheet["A"+index].value = int(answer[0])

		sheet["B"+index].value = answer[2:]
		a = "Present"
		if answers[answer] == False:
			a = "ABSENT"
			redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
			sheet["C"+index].fill = redFill
			sheet["C"+index].alignment = Alignment(horizontal="center")
		sheet["C"+index].value = a
		# filling up stuff in info
		rollnum = int((answer.split(" ")[0]))
		# rollnumbers.append(rollnum)
		name = " ".join(answer.split(" ")[1:])

		info["A"+str(i+5)].value = rollnum
		info["A"+str(i+5)].alignment = Alignment(horizontal="center")
		info["B"+str(i+5)].value = name
		info["B"+str(i+5)].alignment = Alignment(horizontal="center")
		student_money[i+1] = 0
		money_cell = "C"+str(i+5)
		# if info[money_cell].value == None:
		# 	student_money[i+1] = 0
		# else:
		# 	student_money[i+1] = info[money_cell].value
	wb.save(file)

	# dihari = int(int(((info["B1"].value)[3:])/weekdays)/len(answers))
	dihari = int(int(((info["B1"].value)[3:]))/weekdays/len(answers))

	wb.save(file)
	wb = openpyxl.load_workbook(file)



	for i in sheet_names:

		i = wb[i]
		counter = 0
		while isinstance(i["A"+str(4+counter)].value, int):
			# print("counter = " + str(counter))
			if i["C"+str(4+counter)].value == "Present":

				student_money[counter+1] = student_money[counter+1] + dihari
			counter += 1
	print(student_money)
	info = wb["info"]
	for c, i in enumerate(student_money):
		money_cell = "C"+str(c+5)
		# print(student_money[c+1])
		info[money_cell].value = student_money[c+1]


	if int((sheet.title).split(".")[0]) == weekdays:
		print("You have completed " + str(weekdays) + " working days")
		newbookname = str(int((os.listdir("attendance")[-1]).split(".")[0])+1)+".xlsx"
		newbook = openpyxl.Workbook(newbookname)
		wb.save(file)
		newbook.save("attendance\\"+newbookname)
		sys.exit()
	wb.create_sheet(str(int(sheet.title)+1))
	wb.save(file)
	sys.exit()



# attendance
if options.studentattendance:
	attendance()
	load_attendance_into_excel()
# adding student(optional)
if options.student:
	add_one(options.student)
# deleting a student(optional)
if options.removestudent:
	delete_one(options.removestudent)
# know a student
if options.knowone:
	know_one(options.knowone)

