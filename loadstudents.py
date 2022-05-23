import openpyxl
import os
from datetime import datetime
def add_one(s):
	path = "students.xlsx"

	# To open the workbook
	wb = openpyxl.load_workbook(path)

	# Get workbook active sheet object
	sheet = wb.active
	# getting the column
	row = sheet.max_row
	row_a = []

	for i in range(1, row + 1): 
		cell_obj = sheet.cell(row = i, column = 1) 
		row_a.append(cell_obj.value)
	# getting the last element
	last_rn = row_a[-1]
	# make sure it is an integer
	if isinstance(last_rn, int):
		pass
	else:
		for i in row_a[::-1]:
			if isinstance(i, int) == False:
				row -= 1
	# data
	data = [row_a[row]+1, s, "/".join((str(datetime.today())[0:10]).split("-")[::-1])]
	# writing data on the file
	sheet.cell(row= row+2 , column = 1) .value = data[0]
	sheet.cell(row= row+2 , column = 2) .value = data[1]
	sheet.cell(row= row+2 , column = 3) .value = data[2]


	wb.save("students.xlsx")
	print("Roll no. of " + s +" is " + str(data[0]))

def delete_one(rn):
	print("yes")

	wb = openpyxl.load_workbook("students.xlsx")

	sheet =  wb["1"]
	counter = 0
	print("yes")
	while isinstance(sheet["A"+str(counter+2)].value, int):
		if sheet["A"+str(counter+2)].value == int(rn):
			sheet["A"+str(counter+2)].value =None
			sheet["B"+str(counter+2)].value = None
			sheet["C"+str(counter+2)].value = None
		counter += 1
	wb.save("students.xlsx")

# def extract_data(rn, sheet, num):
# 	counter = 0
# 	while isinstance(sheet["A"+str(counter+num)].value, int):
# 		if sheet["A"+str(counter+num)].value == int(rn):
# 			rollnum = sheet["A"+str(counter+num)].value
# 			name = sheet["B"+str(counter+num)].value
# 			date = sheet["C"+str(counter+num)].value
# 			return [rollnum, name, date]
# 		counter += 1

def know_one(rn):
	students = openpyxl.load_workbook("students.xlsx")
	students_sheet =  students["1"]
	total_attendance = 0
	attendance_lst = []
	total_money = 0
	moeny_lst = []
	counter = 0
	while isinstance(students_sheet["A"+str(counter+2)].value, int):
		if students_sheet["A"+str(counter+2)].value == int(rn):
			rollnum = students_sheet["A"+str(counter+2)].value
			name = students_sheet["B"+str(counter+2)].value
			date = students_sheet["C"+str(counter+2)].value
			print(str(rollnum), name, date)
			break
		counter += 1
	# getting attendance info
	file_names = os.listdir("attendance")
	for file in file_names:
		if "lock" in file:
			continue
		monthly_attendance = 0
		monthly_money = 0
		
		file = "attendance\\" + file
		wb = openpyxl.load_workbook(file)
		sheet_names = wb.sheetnames
		for i in sheet_names:
			if i == "info":
				info = wb["info"]
				counter = 0
				while isinstance(info["A"+str(counter+5)].value, int):
					
					if info["A"+str(counter+5)].value == int(rn):
						monthly_money += info["C"+str(counter+5)].value
						total_money +=  info["C"+str(counter+5)].value
					counter += 1
				continue

				
			i = wb[i]
			counter = 0
			while isinstance(i["A"+str(counter+4)].value, int):
				if i["A"+str(counter+4)].value == int(rn):
				# 	rollnum = students_sheet["A"+str(counter+4)].value
				# 	name = students_sheet["B"+str(counter+4)].value
					if i["c"+str(counter+4)].value == "Present":
						total_attendance += 1
						monthly_attendance += 1
				counter += 1
		attendance_lst.append(monthly_attendance)
		moeny_lst.append(monthly_money)
	# printing monthly attendance and money
	for c, i in enumerate(attendance_lst):
		print("\nMonth", str(c+1), "\nattendance =", str(i), "salary =", str(moeny_lst[c]))
	# printing total money and attendace
	print("\nTOTAL ATTENDANCE =", str(total_attendance))
	print("TOTAL SALARY =", str(total_money))
	
			

				